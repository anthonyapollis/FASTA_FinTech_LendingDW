"""
FASTA FinTech Lending DW — ML Results Report Generator
=======================================================
Reads all 4 ML artifact JSONs and writes a formatted Excel + HTML report.
Run: python generate_ml_report.py
"""

import json
import os
from pathlib import Path
from datetime import datetime

HERE      = Path(__file__).parent
ARTIFACTS = HERE / "fabric" / "ml" / "artifacts"
OUT_DIR   = HERE / "reports"
OUT_DIR.mkdir(exist_ok=True)

# ── Load all artifacts ─────────────────────────────────────────────────────────
def load(name):
    path = ARTIFACTS / name
    with open(path) as f:
        return json.load(f)

cr   = load("credit_risk_metrics.json")
aff  = load("affordability_metrics.json")
churn= load("churn_collections_metrics.json")
ch   = load("channel_roi_metrics.json")
pipe = load("pipeline_run_log.json")

with open(HERE / "ml_artifacts" / "ml_summary.json", encoding="utf-8") as f:
    latest = json.load(f)

run_date = datetime.now().strftime("%Y-%m-%d")
latest_credit_auc = latest["model_1_credit_risk"]["auc"]
latest_afford_r2 = latest["model_2_affordability"]["r2"]
latest_afford_mae = latest["model_2_affordability"]["mae_zar"]
latest_default_rate = latest["model_1_credit_risk"]["default_rate_in_test"]

executive_actions = [
    {
        "priority": "P1",
        "theme": "Credit policy",
        "insight": f"The latest portfolio-facing run defaults at {latest_default_rate*100:.1f}%, with high-risk segments materially above portfolio average.",
        "recommendation": "Use model deciles as decision bands: auto-approve low-risk bands, price or limit medium-risk bands, and route high-risk bands to manual review until production AUC clears 0.70.",
        "owner": "Credit Risk",
        "timeframe": "0-30 days",
    },
    {
        "priority": "P1",
        "theme": "Affordability",
        "insight": "The affordability model explains the repayment-capacity signal strongly (R² 0.950), and stress tests show the book is sensitive to income shocks.",
        "recommendation": "Add a stress-tested affordability buffer before approval, especially for lower-income bands and applicants with high debt-service ratios.",
        "owner": "Lending Ops",
        "timeframe": "0-30 days",
    },
    {
        "priority": "P2",
        "theme": "Collections",
        "insight": "Churn and promise-to-pay scores can separate proactive retention from costly late-stage collections.",
        "recommendation": "Create a collections action matrix using churn risk x PTP probability: agent call for high-risk/low-PTP, payment-plan nudge for high-risk/high-PTP, automated SMS for low-risk customers.",
        "owner": "Collections",
        "timeframe": "30-60 days",
    },
    {
        "priority": "P2",
        "theme": "Acquisition ROI",
        "insight": "Email has the best channel quality score and 12.3% default rate; WhatsApp has the weakest quality profile at 28.9% default.",
        "recommendation": "Shift budget toward Email, Direct, and Organic Search; cap WhatsApp/Facebook spend until lead quality improves.",
        "owner": "Marketing",
        "timeframe": "30-60 days",
    },
    {
        "priority": "P1",
        "theme": "MLOps control",
        "insight": "The production value is the operating loop: scheduled scoring, MLflow registration, audit logs, alerting, and retraining when AUC drops.",
        "recommendation": "Keep the weekly model-health gate and add data drift checks before promoting new model versions.",
        "owner": "MLOps",
        "timeframe": "0-30 days",
    },
]

decision_rules = [
    ("Approve faster", "Low-risk deciles with affordability buffer above R2,000", "Instant approve or offer best available rate"),
    ("Price for risk", "Medium-risk deciles or thin affordability buffer", "Lower limit, shorter term, or risk-adjusted price"),
    ("Protect margin", "High-risk deciles, failed affordability, or severe stress sensitivity", "Manual review, decline, or request stronger evidence"),
    ("Recover earlier", "High churn risk before missed-payment spiral", "Proactive call or payment-plan offer before collections escalation"),
    ("Buy better customers", "Channels with lower default and stronger collection rate", "Reallocate acquisition budget to quality, not only volume"),
]

# ── Excel report ───────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter

    NAVY   = "FF1A2E44"
    NAVY_LT = "FF2C3E50"
    WHITE  = "FFFFFFFF"
    GOLD   = "FFFFC107"
    RED    = "FFEE5a6f"
    GREEN  = "FF27ae60"
    GREY   = "FFF5f5f5"

    def hdr(ws, row, col, value, bg=NAVY, fg=WHITE, bold=True, size=11):
        c = ws.cell(row=row, column=col, value=value)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.font      = Font(color=fg, bold=bold, size=size, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        return c

    def cell(ws, row, col, value, bg=None, bold=False, fmt=None, align="center"):
        c = ws.cell(row=row, column=col, value=value)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.font      = Font(bold=bold, size=10, name="Calibri")
        c.alignment = Alignment(horizontal=align, vertical="center")
        if fmt:
            c.number_format = fmt
        return c

    def border_range(ws, min_row, max_row, min_col, max_col):
        thin = Side(style="thin", color="FFCCCCCC")
        for r in range(min_row, max_row+1):
            for c in range(min_col, max_col+1):
                ws.cell(r, c).border = Border(left=thin, right=thin,
                                               top=thin, bottom=thin)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22

    ws.merge_cells("A1:E1")
    hdr(ws, 1, 1, "FASTA FinTech Lending DW — ML Results Summary", size=14)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:E2")
    hdr(ws, 2, 1, f"Run Date: {run_date}  |  Pipeline: {pipe['pipeline']}  |  Duration: {pipe['total_seconds']:.0f}s  |  Status: {pipe['steps_ok']}/{pipe['steps_ok']+pipe['steps_fail']} steps OK",
        bg=NAVY_LT, fg=GOLD, bold=False, size=10)

    row = 4
    for label, val in [
        ("", ""),
        ("MODEL", "KEY METRIC"),
    ]:
        pass

    headers = ["Model", "Algorithm", "Key Metric", "Value", "Status"]
    for i, h in enumerate(headers, 1):
        hdr(ws, 4, i, h)

    rows = [
        ("Credit Risk Classifier",      latest["model_1_credit_risk"]["algorithm"], "AUC", f"{latest_credit_auc:.2f}", "⚠ Low (synthetic data)"),
        ("Affordability Regression",     latest["model_2_affordability"]["algorithm"], "R²", f"{latest_afford_r2:.3f}", "✅ Strong signal"),
        ("Churn + Promise-to-Pay",       "Gradient Boosting",      "Churn AUC",       "1.000", "✅ Excellent"),
        ("Channel ROI Scorecard",        "Quality Scorecard",      "Top Channel",     "EMAIL", "✅ Complete"),
    ]
    for i, (m, a, k, v, s) in enumerate(rows):
        bg = GREY if i % 2 == 0 else WHITE
        cell(ws, 5+i, 1, m, bg=bg, bold=True, align="left")
        cell(ws, 5+i, 2, a, bg=bg)
        cell(ws, 5+i, 3, k, bg=bg)
        cell(ws, 5+i, 4, v, bg=bg, bold=True)
        cell(ws, 5+i, 5, s, bg=bg)
    border_range(ws, 4, 8, 1, 5)

    row = 10
    ws.merge_cells(f"A{row}:E{row}")
    hdr(ws, row, 1, "Pipeline Execution Log", bg=NAVY_LT, fg=GOLD)
    row += 1
    for h, col in zip(["Notebook","Label","Status","Elapsed (sec)","Run At"], range(1,6)):
        hdr(ws, row, col, h, size=10)
    row += 1
    for i, nb in enumerate(pipe["notebook_runs"]):
        bg = GREY if i % 2 == 0 else WHITE
        cell(ws, row, 1, nb["notebook"],    bg=bg, align="left")
        cell(ws, row, 2, nb["label"],       bg=bg, align="left")
        c = cell(ws, row, 3, nb["status"],  bg=bg)
        c.font = Font(color=GREEN if nb["status"]=="SUCCESS" else RED, bold=True, size=10)
        cell(ws, row, 4, nb["elapsed_sec"], bg=bg, fmt="0.00")
        cell(ws, row, 5, nb["run_at"],      bg=bg)
        row += 1
    border_range(ws, 11, row-1, 1, 5)

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    hdr(ws, row, 1, "Executive Action Plan", bg=NAVY_LT, fg=GOLD)
    row += 1
    for h, col in zip(["Priority", "Theme", "Insight", "Recommendation", "Owner / Timing"], range(1,6)):
        hdr(ws, row, col, h, size=10)
    row += 1
    start_row = row
    for i, action in enumerate(executive_actions):
        bg = GREY if i % 2 == 0 else WHITE
        cell(ws, row, 1, action["priority"], bg=bg, bold=True)
        cell(ws, row, 2, action["theme"], bg=bg, bold=True, align="left")
        cell(ws, row, 3, action["insight"], bg=bg, align="left")
        cell(ws, row, 4, action["recommendation"], bg=bg, align="left")
        cell(ws, row, 5, f"{action['owner']} | {action['timeframe']}", bg=bg, align="left")
        ws.row_dimensions[row].height = 48
        row += 1
    border_range(ws, start_row-1, row-1, 1, 5)

    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    hdr(ws, row, 1, "Decision Rules for FASTA", bg=NAVY_LT, fg=GOLD)
    row += 1
    for h, col in zip(["Decision", "Trigger", "Action"], range(1,4)):
        hdr(ws, row, col, h, size=10)
    row += 1
    start_row = row
    for i, (decision, trigger, action) in enumerate(decision_rules):
        bg = GREY if i % 2 == 0 else WHITE
        cell(ws, row, 1, decision, bg=bg, bold=True, align="left")
        cell(ws, row, 2, trigger, bg=bg, align="left")
        cell(ws, row, 3, action, bg=bg, align="left")
        ws.row_dimensions[row].height = 34
        row += 1
    border_range(ws, start_row-1, row-1, 1, 3)

    # ── Sheet 2: Credit Risk ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Credit Risk")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 20
    for col in "BCDEF": ws2.column_dimensions[col].width = 18

    ws2.merge_cells("A1:F1")
    hdr(ws2, 1, 1, "Model 1 — Credit Risk Classifier", size=13)
    ws2.row_dimensions[1].height = 28

    meta = [
        ("Best Model", cr["best_model"]),
        ("Train Size", cr["train_size"]),
        ("Test Size",  cr["test_size"]),
        ("Default Rate", f"{cr['default_rate_overall']*100:.1f}%"),
    ]
    for i, (k, v) in enumerate(meta):
        cell(ws2, 3+i, 1, k, bold=True, align="left")
        cell(ws2, 3+i, 2, str(v), align="left")

    row = 8
    ws2.merge_cells(f"A{row}:F{row}")
    hdr(ws2, row, 1, "Model Comparison", bg=NAVY_LT, fg=GOLD)
    row += 1
    for h, col in zip(["Model","AUC","Avg Precision","Brier Score","CV-AUC Mean","CV-AUC Std"], range(1,7)):
        hdr(ws2, row, col, h, size=10)
    row += 1
    for i, (name, m) in enumerate(cr["models"].items()):
        bg = GOLD if name == cr["best_model"] else (GREY if i%2==0 else WHITE)
        bold = name == cr["best_model"]
        cell(ws2, row, 1, name,               bg=bg, bold=bold, align="left")
        cell(ws2, row, 2, m["auc"],           bg=bg, bold=bold, fmt="0.000")
        cell(ws2, row, 3, m["avg_precision"], bg=bg, fmt="0.000")
        cell(ws2, row, 4, m["brier_score"],   bg=bg, fmt="0.000")
        cell(ws2, row, 5, m["cv_auc_mean"],   bg=bg, bold=bold, fmt="0.000")
        cell(ws2, row, 6, m["cv_auc_std"],    bg=bg, fmt="0.000")
        row += 1
    border_range(ws2, 9, row-1, 1, 6)

    row += 1
    ws2.merge_cells(f"A{row}:F{row}")
    hdr(ws2, row, 1, "Risk Decile Table", bg=NAVY_LT, fg=GOLD)
    row += 1
    for h, col in zip(["Decile","N Loans","Defaults","Default Rate","Avg Prob"], range(1,6)):
        hdr(ws2, row, col, h, size=10)
    row += 1
    for i, d in enumerate(cr["risk_decile_table"]):
        bg = GREY if i%2==0 else WHITE
        cell(ws2, row, 1, d["risk_decile"],  bg=bg, bold=True)
        cell(ws2, row, 2, d["n"],            bg=bg)
        cell(ws2, row, 3, d["defaults"],     bg=bg)
        c = cell(ws2, row, 4, d["default_rate"], bg=bg, fmt="0.0%")
        if d["default_rate"] > 0.30:
            c.font = Font(color=RED, bold=True, size=10)
        cell(ws2, row, 5, round(d["avg_prob"],4), bg=bg, fmt="0.0000")
        row += 1
    border_range(ws2, row-10, row-1, 1, 5)

    # ── Sheet 3: Affordability ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Affordability")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 28
    for col in "BCDE": ws3.column_dimensions[col].width = 18

    ws3.merge_cells("A1:E1")
    hdr(ws3, 1, 1, "Model 2 — Affordability Regression + Stress Test", size=13)
    ws3.row_dimensions[1].height = 28

    row = 3
    for h, col in zip(["Model","R²","MAE (ZAR)","RMSE","CV-R² Mean"], range(1,6)):
        hdr(ws3, row, col, h, size=10)
    row += 1
    for i, (name, m) in enumerate(aff["models"].items()):
        bg = GOLD if name == aff["best_model"] else (GREY if i%2==0 else WHITE)
        bold = name == aff["best_model"]
        cell(ws3, row, 1, name,          bg=bg, bold=bold, align="left")
        cell(ws3, row, 2, m["r2"],       bg=bg, bold=bold, fmt="0.000")
        cell(ws3, row, 3, m["mae"],      bg=bg, fmt='R#,##0')
        cell(ws3, row, 4, m["rmse"],     bg=bg, fmt='R#,##0')
        cell(ws3, row, 5, m["cv_r2_mean"], bg=bg, bold=bold, fmt="0.000")
        row += 1
    border_range(ws3, 3, row-1, 1, 5)

    row += 1
    ws3.merge_cells(f"A{row}:E{row}")
    hdr(ws3, row, 1, "Stress Test Scenarios", bg=NAVY_LT, fg=GOLD)
    row += 1
    for h, col in zip(["Scenario","Avg Affordability","% Unaffordable","% >R2K Available"], range(1,5)):
        hdr(ws3, row, col, h, size=10)
    row += 1
    for i, s in enumerate(aff["stress_scenarios"]):
        bg = GREY if i%2==0 else WHITE
        cell(ws3, row, 1, s["Scenario"],      bg=bg, bold=i==0, align="left")
        cell(ws3, row, 2, s["Avg Affordability"], bg=bg, bold=i==0)
        pct_unaff = s["% Unaffordable"]
        c = cell(ws3, row, 3, pct_unaff, bg=bg)
        try:
            val = float(pct_unaff.replace("%",""))
            if val > 50: c.font = Font(color=RED, bold=True, size=10)
            elif val > 30: c.font = Font(color="FFFF6600", bold=False, size=10)
        except: pass
        cell(ws3, row, 4, s["% >R2K available"], bg=bg)
        row += 1
    border_range(ws3, row-len(aff["stress_scenarios"])-1, row-1, 1, 4)

    # ── Sheet 4: Churn ────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Churn & Collections")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 30
    for col in "BCDE": ws4.column_dimensions[col].width = 18

    ws4.merge_cells("A1:E1")
    hdr(ws4, 1, 1, "Model 3 — Churn + Promise-to-Pay Classifier", size=13)
    ws4.row_dimensions[1].height = 28

    row = 3
    churn_meta = [
        ("Definition",    churn["churn_model"]["definition"]),
        ("Dataset Size",  churn["churn_model"]["dataset_size"]),
        ("Churn Rate",    f"{churn['churn_model']['churn_rate']*100:.1f}%"),
        ("Best Model",    churn["churn_model"]["best_model"]),
    ]
    for k, v in churn_meta:
        cell(ws4, row, 1, k, bold=True, align="left")
        cell(ws4, row, 2, str(v), align="left")
        row += 1

    row += 1
    for h, col in zip(["Model","AUC","Avg Precision"], range(1,4)):
        hdr(ws4, row, col, h, size=10)
    row += 1
    for i, (name, m) in enumerate(churn["churn_model"]["models"].items()):
        bg = GOLD if name == churn["churn_model"]["best_model"] else (GREY if i%2==0 else WHITE)
        cell(ws4, row, 1, name,    bg=bg, bold=True, align="left")
        cell(ws4, row, 2, m["auc"],bg=bg, bold=True, fmt="0.000")
        cell(ws4, row, 3, m["ap"], bg=bg, fmt="0.000")
        row += 1
    border_range(ws4, row-3, row-1, 1, 3)

    row += 1
    ws4.merge_cells(f"A{row}:E{row}")
    hdr(ws4, row, 1, "Promise-to-Pay Model", bg=NAVY_LT, fg=GOLD)
    row += 1
    ptp_meta = [
        ("Definition", churn["ptp_model"]["definition"]),
        ("Dataset Size", churn["ptp_model"]["dataset_size"]),
        ("PTP Rate",  f"{churn['ptp_model']['ptp_rate']*100:.1f}%"),
        ("AUC",       churn["ptp_model"]["auc"]),
    ]
    for k, v in ptp_meta:
        cell(ws4, row, 1, k, bold=True, align="left")
        cell(ws4, row, 2, str(v), align="left")
        row += 1

    # ── Sheet 5: Channel ROI ──────────────────────────────────────────────────
    ws5 = wb.create_sheet("Channel ROI")
    ws5.sheet_view.showGridLines = False
    ws5.column_dimensions["A"].width = 8
    ws5.column_dimensions["B"].width = 26
    for col in "CDEFGH": ws5.column_dimensions[col].width = 18

    ws5.merge_cells("A1:H1")
    hdr(ws5, 1, 1, f"Model 4 — Channel ROI Scorecard  |  Total Budget: R{ch['total_budget_optimised']:,}", size=13)
    ws5.row_dimensions[1].height = 28

    row = 3
    for h, col in zip(["Rank","Channel","N Loans","Default Rate","Quality Score",
                        "Avg Collection","Avg Credit Score","Budget (ZAR)"], range(1,9)):
        hdr(ws5, row, col, h, size=10)
    row += 1
    for i, c_ in enumerate(ch["channel_quality_table"]):
        bg = GOLD if i==0 else (GREY if i%2==0 else WHITE)
        bold = i==0
        cell(ws5, row, 1, c_["rank"],                    bg=bg, bold=bold)
        cell(ws5, row, 2, c_["channel_std"],             bg=bg, bold=bold, align="left")
        cell(ws5, row, 3, c_["n_loans"],                 bg=bg)
        c_cell = cell(ws5, row, 4, c_["default_rate"],   bg=bg, fmt="0.0%")
        if c_["default_rate"] < 0.15:
            c_cell.font = Font(color=GREEN, bold=True, size=10)
        elif c_["default_rate"] > 0.25:
            c_cell.font = Font(color=RED, bold=False, size=10)
        cell(ws5, row, 5, c_["quality_score"],           bg=bg, bold=bold, fmt="0.000")
        cell(ws5, row, 6, c_["avg_collection"],          bg=bg, fmt="0.000")
        cell(ws5, row, 7, c_["avg_credit_score"],        bg=bg, fmt="0.0")
        cell(ws5, row, 8, c_["recommended_budget_zar"],  bg=bg, bold=bold, fmt='R#,##0')
        row += 1
    border_range(ws5, 3, row-1, 1, 8)

    excel_path = OUT_DIR / f"FASTA_ML_Results_{run_date}.xlsx"
    try:
        wb.save(excel_path)
    except PermissionError:
        excel_path = OUT_DIR / f"FASTA_ML_Results_{run_date}_{datetime.now():%H%M%S}.xlsx"
        wb.save(excel_path)
        print("Dated Excel report was locked; saved a timestamped copy instead.")
    print(f"Excel saved: {excel_path}")

except ImportError:
    print("openpyxl not installed — skipping Excel. Run: pip install openpyxl")
    excel_path = None

# ── HTML report ────────────────────────────────────────────────────────────────
html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>FASTA ML Results — {run_date}</title>
<style>
  body {{ font-family: Calibri, Arial, sans-serif; background:#eaf3f5; color:#172b33; margin:0; padding:24px }}
  h1 {{ background:linear-gradient(135deg,#006272 0%,#008d92 70%); color:#fff; padding:24px 28px; border-radius:8px; margin-bottom:8px; border-bottom:5px solid #ffc107 }}
  h2 {{ background:#d9f0f2; color:#006272; padding:12px 16px; border-radius:6px; margin:26px 0 10px; border-left:6px solid #ffc107 }}
  h3 {{ color:#006272; margin:16px 0 4px }}
  table {{ border-collapse:collapse; width:100%; margin-bottom:18px; background:#fff; border-radius:8px; overflow:hidden; box-shadow:0 4px 14px rgba(0,98,114,.12) }}
  th {{ background:#006272; color:#fff; padding:10px 12px; text-align:center; font-size:13px }}
  td {{ padding:9px 12px; text-align:center; font-size:13px; border-bottom:1px solid #d9e7ea }}
  tr:nth-child(even) td {{ background:#f3fafb }}
  .best {{ background:#fff1b8 !important; font-weight:bold }}
  .good {{ color:#00995c; font-weight:bold }}
  .warn {{ color:#e89b00; font-weight:bold }}
  .bad  {{ color:#d94d55; font-weight:bold }}
  .meta {{ display:flex; gap:20px; flex-wrap:wrap; margin-bottom:16px }}
  .meta-card {{ background:#fff; border:1px solid #c7dee3; border-top:5px solid #ffc107; border-radius:8px; padding:16px 22px; min-width:180px; box-shadow:0 4px 12px rgba(0,0,0,.08) }}
  .meta-card:nth-child(2) {{ border-top-color:#27ae60 }}
  .meta-card:nth-child(3) {{ border-top-color:#00bcd4 }}
  .meta-card:nth-child(4) {{ border-top-color:#ee5a6f }}
  .meta-card:nth-child(5) {{ border-top-color:#27ae60 }}
  .meta-card:nth-child(6) {{ border-top-color:#ff7043 }}
  .meta-card .val {{ font-size:25px; font-weight:bold; color:#006272 }}
  .meta-card .lbl {{ font-size:12px; color:#5b6f76; margin-top:4px }}
  .subtitle {{ color:#555; font-size:14px; margin:0 0 20px }}
  .story {{ background:#fffdf3; border:1px solid #ffc107; border-radius:8px; padding:20px 24px; margin:20px 0; box-shadow:0 4px 14px rgba(255,193,7,.18) }}
  .story h2 {{ margin-top:0 }}
  .action-grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(240px,1fr)); gap:12px; margin:14px 0 22px }}
  .action-card {{ background:#fff; border-left:6px solid #006272; border-radius:8px; padding:16px 18px; box-shadow:0 4px 12px rgba(0,0,0,.08) }}
  .action-card:nth-child(2) {{ border-left-color:#27ae60 }}
  .action-card:nth-child(3) {{ border-left-color:#00bcd4 }}
  .action-card:nth-child(4) {{ border-left-color:#ff7043 }}
  .action-card:nth-child(5) {{ border-left-color:#ee5a6f }}
  .action-card .prio {{ color:#006272; font-weight:bold; font-size:12px; text-transform:uppercase }}
  .action-card .theme {{ font-weight:bold; margin:4px 0 6px }}
  .action-card p {{ margin:0; font-size:13px; line-height:1.45 }}
  .left {{ text-align:left }}
</style>
</head>
<body>
<h1>FASTA FinTech Lending DW — ML Results</h1>
<p class="subtitle">Run: {run_date} &nbsp;|&nbsp; Pipeline: {pipe['pipeline']} &nbsp;|&nbsp;
Duration: {pipe['total_seconds']:.0f}s &nbsp;|&nbsp;
Steps: {pipe['steps_ok']}/{pipe['steps_ok']+pipe['steps_fail']} OK</p>

<div class="meta">
  <div class="meta-card"><div class="val">{latest_credit_auc:.2f}</div><div class="lbl">Credit Risk AUC</div></div>
  <div class="meta-card"><div class="val">{latest_afford_r2:.3f}</div><div class="lbl">Affordability R²</div></div>
  <div class="meta-card"><div class="val">R{latest_afford_mae:,.0f}</div><div class="lbl">Affordability MAE</div></div>
  <div class="meta-card"><div class="val">1.000</div><div class="lbl">Churn AUC</div></div>
  <div class="meta-card"><div class="val">EMAIL</div><div class="lbl">Top Channel</div></div>
  <div class="meta-card"><div class="val">R500K</div><div class="lbl">Budget Optimised</div></div>
</div>

<div class="story">
<h2>Executive Story — What FASTA Should Do Next</h2>
<p>This report is not only a model scorecard. It is an operating plan for a lender: approve the right customers faster, protect margin under affordability stress, intervene earlier in collections, and spend marketing budget on channels that bring borrowers who repay.</p>
<p><strong>Important context:</strong> the source data is synthetic, so credit-risk AUC is not presented as a production performance claim. The value demonstrated here is the full production pattern: governed data layers, repeatable scoring, MLflow registration, audit logging, model-health gates, and business-facing actions.</p>
</div>

<h2>Recommended Actions</h2>
<div class="action-grid">
"""
for action in executive_actions:
    html += f"""  <div class="action-card">
    <div class="prio">{action['priority']} · {action['owner']} · {action['timeframe']}</div>
    <div class="theme">{action['theme']}</div>
    <p><strong>Insight:</strong> {action['insight']}</p>
    <p><strong>Recommendation:</strong> {action['recommendation']}</p>
  </div>
"""

html += """</div>

<h2>Decision Rules</h2>
<table>
<tr><th>Decision</th><th>Trigger</th><th>Action</th></tr>
"""
for decision, trigger, action in decision_rules:
    html += f'<tr><td class="left"><strong>{decision}</strong></td><td class="left">{trigger}</td><td class="left">{action}</td></tr>\n'

html += f"""
</table>

<h2>Model 1 — Credit Risk Classifier</h2>
<p><strong>Headline metric:</strong> latest dashboard AUC is <strong>{latest_credit_auc:.2f}</strong>. The table below is the Fabric validation diagnostic table, including cross-validation metrics from the retrain notebook.</p>
<table>
<tr><th>Model</th><th>AUC</th><th>Avg Precision</th><th>Brier Score</th><th>CV-AUC Mean</th><th>CV-AUC Std</th></tr>
"""
for name, m in cr["models"].items():
    cls = 'class="best"' if name == cr["best_model"] else ""
    html += f'<tr {cls}><td>{name}</td><td>{m["auc"]:.3f}</td><td>{m["avg_precision"]:.3f}</td><td>{m["brier_score"]:.3f}</td><td>{m["cv_auc_mean"]:.3f}</td><td>{m["cv_auc_std"]:.3f}</td></tr>\n'

html += """</table>
<h3>Risk Decile Table</h3>
<table>
<tr><th>Decile</th><th>N Loans</th><th>Defaults</th><th>Default Rate</th><th>Avg Prob</th></tr>
"""
for d in cr["risk_decile_table"]:
    dr = d["default_rate"]
    cls = 'class="bad"' if dr > 0.30 else ('class="warn"' if dr > 0.20 else 'class="good"')
    html += f'<tr><td>{d["risk_decile"]}</td><td>{d["n"]}</td><td>{d["defaults"]}</td><td {cls}>{dr:.1%}</td><td>{d["avg_prob"]:.4f}</td></tr>\n'

html += f"""</table>

<h2>Model 2 — Affordability Regression</h2>
<p><strong>Headline metric:</strong> latest dashboard R² is <strong>{latest_afford_r2:.3f}</strong>. The table below is the Fabric validation diagnostic table from the retrain notebook.</p>
<table>
<tr><th>Model</th><th>R²</th><th>MAE (ZAR)</th><th>RMSE</th><th>CV-R² Mean</th></tr>
"""
for name, m in aff["models"].items():
    cls = 'class="best"' if name == aff["best_model"] else ""
    html += f'<tr {cls}><td>{name}</td><td>{m["r2"]:.3f}</td><td>R{m["mae"]:,.0f}</td><td>R{m["rmse"]:,.0f}</td><td>{m["cv_r2_mean"]:.3f}</td></tr>\n'

html += "<h3>Stress Test Scenarios</h3><table><tr><th>Scenario</th><th>Avg Affordability</th><th>% Unaffordable</th><th>% >R2K Available</th></tr>\n"
for s in aff["stress_scenarios"]:
    try:
        pct = float(s["% Unaffordable"].replace("%",""))
        cls = 'class="bad"' if pct > 50 else ('class="warn"' if pct > 30 else 'class="good"')
    except: cls = ""
    html += f'<tr><td>{s["Scenario"]}</td><td>{s["Avg Affordability"]}</td><td {cls}>{s["% Unaffordable"]}</td><td>{s["% >R2K available"]}</td></tr>\n'

html += f"""</table>

<h2>Model 3 — Churn + Promise-to-Pay</h2>
<table>
<tr><th>Model</th><th>Churn AUC</th><th>Avg Precision</th></tr>
"""
for name, m in churn["churn_model"]["models"].items():
    cls = 'class="best"' if name == churn["churn_model"]["best_model"] else ""
    html += f'<tr {cls}><td>{name}</td><td>{m["auc"]:.3f}</td><td>{m["ap"]:.3f}</td></tr>\n'
html += f'</table><p>PTP AUC: <strong>{churn["ptp_model"]["auc"]}</strong> | Churn Rate: <strong>{churn["churn_model"]["churn_rate"]*100:.1f}%</strong> | PTP Rate: <strong>{churn["ptp_model"]["ptp_rate"]*100:.1f}%</strong></p>'

html += f"""
<h2>Model 4 — Channel ROI Scorecard (Total Budget: R{ch['total_budget_optimised']:,})</h2>
<table>
<tr><th>Rank</th><th>Channel</th><th>N Loans</th><th>Default Rate</th><th>Quality Score</th><th>Avg Collection</th><th>Avg Credit Score</th><th>Budget (ZAR)</th></tr>
"""
for i, c_ in enumerate(ch["channel_quality_table"]):
    cls = 'class="best"' if i == 0 else ""
    dr = c_["default_rate"]
    dr_cls = "good" if dr < 0.15 else ("bad" if dr > 0.25 else "")
    html += f'<tr {cls}><td>{c_["rank"]}</td><td>{c_["channel_std"]}</td><td>{c_["n_loans"]}</td><td class="{dr_cls}">{dr:.1%}</td><td>{c_["quality_score"]:.3f}</td><td>{c_["avg_collection"]:.3f}</td><td>{c_["avg_credit_score"]:.0f}</td><td>R{c_["recommended_budget_zar"]:,.0f}</td></tr>\n'

html += f"""</table>
<p style="color:#888;font-size:12px;margin-top:40px">Generated {run_date} | FASTA FinTech LendingDW | Anthony Apollis</p>
</body></html>"""

html_path = OUT_DIR / f"FASTA_ML_Results_{run_date}.html"
html_path.write_text(html, encoding="utf-8")
print(f"HTML saved: {html_path}")
print(f"\nOpen in browser: {html_path}")
